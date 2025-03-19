import boto3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta


# Base class for cost reports
class CostReport:
    def __init__(self, output_file):
        self.client = boto3.client('ce', region_name='ca-central-1')
        self.output_file = output_file

    def fetch_cost_data(self, time_period, group_by):
        response = self.client.get_cost_and_usage(
            TimePeriod=time_period,
            Granularity='MONTHLY',
            Metrics=['UnblendedCost'],
            GroupBy=group_by
        )
        return response['ResultsByTime'][0]['Groups']

    def generate_report(self):
        raise NotImplementedError("Subclasses must implement this method")


# Projectwise report class
class ProjectServiceReport(CostReport):
    def get_time_periods(self):
        current_month = datetime.now().replace(day=1)
        previous_month = (current_month - timedelta(days=1)).replace(day=1)
        time_period_current = {
            'Start': current_month.strftime('%Y-%m-%d'),
            'End': (current_month.replace(month=current_month.month % 12 + 1)).strftime('%Y-%m-%d')
        }
        time_period_previous = {
            'Start': previous_month.strftime('%Y-%m-%d'),
            'End': (previous_month.replace(month=previous_month.month % 12 + 1)).strftime('%Y-%m-%d')
        }
        return time_period_current, time_period_previous

    def process_data(self, cost_data_current, cost_data_previous):
        project_costs = {}
        for item in cost_data_current:
            project_tag = next((tag.split('$')[1] for tag in item['Keys'] if tag.startswith('Project$')),
                               'No Project Tag')
            service = next((svc for svc in item['Keys'] if not svc.startswith('Project$')), 'No Service')
            amount = float(item['Metrics']['UnblendedCost']['Amount'])
            if project_tag not in project_costs:
                project_costs[project_tag] = {}
            project_costs[project_tag][service] = {'Current Cost': amount}

        for item in cost_data_previous:
            project_tag = next((tag.split('$')[1] for tag in item['Keys'] if tag.startswith('Project$')),
                               'No Project Tag')
            service = next((svc for svc in item['Keys'] if not svc.startswith('Project$')), 'No Service')
            amount = float(item['Metrics']['UnblendedCost']['Amount'])
            if project_tag not in project_costs:
                project_costs[project_tag] = {}
            if service not in project_costs[project_tag]:
                project_costs[project_tag][service] = {}
            project_costs[project_tag][service]['Previous Cost'] = amount

        return project_costs

    def generate_excel(self, project_costs, time_period_previous, time_period_current):
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            used_sheet_names = set()
            for project, services in project_costs.items():
                sheet_name = self.get_unique_sheet_name(project, used_sheet_names)
                used_sheet_names.add(sheet_name)
                data = []
                for service, costs in services.items():
                    current_cost = costs.get('Current Cost', 0)
                    previous_cost = costs.get('Previous Cost', 0)
                    difference = current_cost - previous_cost
                    data.append({
                        'Service': service,
                        'Current Cost (USD)': round(current_cost, 2),
                        'Previous Cost (USD)': round(previous_cost, 2),
                        'Difference (USD)': round(difference, 2)
                    })
                df = pd.DataFrame(data)
                df = df.sort_values(by='Current Cost (USD)', ascending=False)
                total_current_cost = df['Current Cost (USD)'].sum()
                total_previous_cost = df['Previous Cost (USD)'].sum()
                total_difference = total_current_cost - total_previous_cost
                total_row = pd.DataFrame({
                    'Service': ['Total'],
                    'Current Cost (USD)': [round(total_current_cost, 2)],
                    'Previous Cost (USD)': [round(total_previous_cost, 2)],
                    'Difference (USD)': [round(total_difference, 2)]
                })
                df = pd.concat([df, total_row], ignore_index=True)
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)

        workbook = load_workbook(self.output_file)
        light_yellow = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
        light_cyan = PatternFill(start_color='E0FFFF', end_color='E0FFFF', fill_type='solid')
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            title = f"Report Period: {time_period_previous['Start']} to {time_period_previous['End']} & " \
                    f"{time_period_current['Start']} to {time_period_current['End']}"
            sheet.cell(row=1, column=1, value=title).fill = light_yellow
            sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=1, column=1).border = thin_border
            sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
            last_row = sheet.max_row
            for row in sheet.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=4):
                for cell in row:
                    if cell.column == 4:  # Difference column
                        try:
                            value = float(cell.value or 0)
                            cell.fill = red_fill if value > 0 else green_fill if value < 0 else light_cyan
                        except (ValueError, TypeError):
                            cell.fill = light_cyan
                    else:
                        cell.fill = light_yellow if cell.column == 1 else light_cyan
                    cell.border = thin_border
            for col in range(1, 5):
                sheet.column_dimensions[get_column_letter(col)].width = 30

        workbook.save(self.output_file)
        print(f"Consolidated report saved as {self.output_file}")

    def generate_report(self):
        time_period_current, time_period_previous = self.get_time_periods()
        cost_data_current = self.fetch_cost_data(time_period_current, [{'Type': 'TAG', 'Key': 'Project'},
                                                                       {'Type': 'DIMENSION', 'Key': 'SERVICE'}])
        cost_data_previous = self.fetch_cost_data(time_period_previous, [{'Type': 'TAG', 'Key': 'Project'},
                                                                         {'Type': 'DIMENSION', 'Key': 'SERVICE'}])
        project_costs = self.process_data(cost_data_current, cost_data_previous)
        self.generate_excel(project_costs, time_period_previous, time_period_current)

    @staticmethod
    def get_unique_sheet_name(project, used_names):
        base_name = project[:28] if project.strip() else 'Unnamed_Project'
        sheet_name = base_name
        counter = 1
        while sheet_name.lower() in [name.lower() for name in used_names]:
            sheet_name = f"{base_name}_{counter}"
            counter += 1
        return sheet_name


# AWS Account total cost report class
class TotalCostsReport(CostReport):
    def get_date_ranges(self):
        today = datetime.today()
        current_month_start = datetime(today.year, today.month, 1)
        first_month_start = current_month_start - relativedelta(months=1)
        first_month_end = current_month_start - timedelta(days=1)
        second_month_start = current_month_start - relativedelta(months=2)
        second_month_end = first_month_start - timedelta(days=1)
        return {
            "first_month": {
                "start_date": second_month_start.strftime('%Y-%m-%d'),
                "end_date": second_month_end.strftime('%Y-%m-%d'),
            },
            "second_month": {
                "start_date": first_month_start.strftime('%Y-%m-%d'),
                "end_date": first_month_end.strftime('%Y-%m-%d'),
            },
        }

    def process_data(self, first_month_data, second_month_data):
        first_month_costs = {
            item['Keys'][0].split('$')[1] if item['Keys'][0].startswith('Project$') else 'No Project Tag':
                float(item['Metrics']['UnblendedCost']['Amount']) for item in first_month_data}
        second_month_costs = {
            item['Keys'][0].split('$')[1] if item['Keys'][0].startswith('Project$') else 'No Project Tag':
                float(item['Metrics']['UnblendedCost']['Amount']) for item in second_month_data}

        all_projects = set(first_month_costs.keys()).union(second_month_costs.keys())
        data = []
        for project in all_projects:
            first_cost = first_month_costs.get(project, 0)
            second_cost = second_month_costs.get(project, 0)
            data.append({
                'Project': project,
                f"Cost of {self.second_month['start_date']} to {self.second_month['end_date']}": round(second_cost, 2),
                f"Cost of {self.first_month['start_date']} to {self.first_month['end_date']}": round(first_cost, 2),
                'Difference': round(first_cost - second_cost, 2)
            })
        df = pd.DataFrame(data)
        total_first = df[f"Cost of {self.first_month['start_date']} to {self.first_month['end_date']}"].sum()
        total_second = df[f"Cost of {self.second_month['start_date']} to {self.second_month['end_date']}"].sum()
        total_difference = total_first - total_second
        totals_row = pd.DataFrame({
            'Project': ['Total'],
            f"Cost of {self.second_month['start_date']} to {self.second_month['end_date']}": [round(total_second, 2)],
            f"Cost of {self.first_month['start_date']} to {self.first_month['end_date']}": [round(total_first, 2)],
            'Difference': [round(total_difference, 2)]
        })
        return pd.concat([df, totals_row], ignore_index=True)

    def generate_excel(self, merged_df):
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            merged_df.to_excel(writer, index=False, sheet_name='Report', startrow=2, startcol=0)

        wb = load_workbook(self.output_file)
        sheet = wb['Report']
        first_month_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
        second_month_fill = PatternFill(start_color='E0FFFF', end_color='E0FFFF', fill_type='solid')
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        sheet.cell(row=1,
                   column=1).value = f"Report Period: {self.first_month['start_date']} to {self.first_month['end_date']} & " \
                                     f"{self.second_month['start_date']} to {self.second_month['end_date']}"
        sheet.cell(row=1, column=1).fill = first_month_fill
        sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=1, column=1).border = thin_border
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

        last_row = sheet.max_row
        for row in sheet.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=4):
            for cell in row:
                if cell.column == 4:  # Difference column
                    try:
                        difference = float(cell.value or 0)
                        cell.fill = red_fill if difference > 0 else green_fill if difference < 0 else second_month_fill
                    except (ValueError, TypeError):
                        cell.fill = second_month_fill
                else:
                    cell.fill = first_month_fill if cell.column == 1 else second_month_fill
                cell.border = thin_border
        for col in range(1, 5):
            sheet.column_dimensions[get_column_letter(col)].width = 30

        wb.save(self.output_file)
        print(f"Report saved as {self.output_file}")

    def generate_report(self):
        date_ranges = self.get_date_ranges()
        self.first_month = date_ranges["first_month"]
        self.second_month = date_ranges["second_month"]
        first_month_data = self.fetch_cost_data({
            'Start': self.first_month['start_date'],
            'End': self.first_month['end_date']
        }, [{'Type': 'TAG', 'Key': 'Project'}])
        second_month_data = self.fetch_cost_data({
            'Start': self.second_month['start_date'],
            'End': self.second_month['end_date']
        }, [{'Type': 'TAG', 'Key': 'Project'}])
        merged_df = self.process_data(first_month_data, second_month_data)
        self.generate_excel(merged_df)


# Display the menu
def display_menu():
    print("\nSelect a cost report type:")
    print("1. AWS Account total cost report")
    print("2. Projectwise cost report")
    print("3. Exit")


# Main function to handle user interaction
def main():
    while True:
        display_menu()
        choice = input("Pick an option (1-3): ")
        if choice == '1':
            output_file = 'AWS_Project_Cost_Report_with_Difference_and_Totals.xlsx'
            report = TotalCostsReport(output_file)
            report.generate_report()
        elif choice == '2':
            output_file = 'Projectwise_Cost_Report.xlsx'
            report = ProjectServiceReport(output_file)
            report.generate_report()
        elif choice == '3':
            print("Exiting the program.")
            break
        else:
            print("[Invalid Selection] select an option from the list.")


# Run the program
if __name__ == "__main__":
    main()
