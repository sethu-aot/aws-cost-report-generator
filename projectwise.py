import boto3
import pandas as pd
import datetime
import xlsxwriter
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Initialize the AWS Cost Explorer client
client = boto3.client('ce', region_name='ca-central-1')

# Define the time period for current and previous month
current_month = datetime.datetime.now().replace(day=1)
previous_month = (current_month - datetime.timedelta(days=1)).replace(day=1)

time_period_current = {
    'Start': current_month.strftime('%Y-%m-%d'),
    'End': (current_month.replace(month=current_month.month % 12 + 1)).strftime('%Y-%m-%d')
}

time_period_previous = {
    'Start': previous_month.strftime('%Y-%m-%d'),
    'End': (previous_month.replace(month=previous_month.month % 12 + 1)).strftime('%Y-%m-%d')
}


def get_unique_sheet_name(project, used_names):
    """Generate a unique sheet name handling case sensitivity and duplicates."""
    base_name = project[:28] if project.strip() else 'Unnamed_Project'
    sheet_name = base_name
    counter = 1

    while sheet_name.lower() in [name.lower() for name in used_names]:
        sheet_name = f"{base_name}_{counter}"
        counter += 1

    return sheet_name


def fetch_cost_data(time_period):
    """Fetch AWS cost data grouped by project and service."""
    response = client.get_cost_and_usage(
        TimePeriod=time_period,
        Granularity='MONTHLY',
        Metrics=['UnblendedCost'],
        GroupBy=[{'Type': 'TAG', 'Key': 'Project'}, {'Type': 'DIMENSION', 'Key': 'SERVICE'}]
    )
    return response['ResultsByTime'][0]['Groups']


# Fetch data for both months
cost_data_current = fetch_cost_data(time_period_current)
cost_data_previous = fetch_cost_data(time_period_previous)

# Dictionary to hold data for each project
project_costs = {}

# Process the current month data
for item in cost_data_current:
    project_tag = next((tag.split('$')[1] for tag in item['Keys'] if tag.startswith('Project$')), 'No Project Tag')
    service = next((svc for svc in item['Keys'] if not svc.startswith('Project$')), 'No Service')
    amount = float(item['Metrics']['UnblendedCost']['Amount'])

    if project_tag not in project_costs:
        project_costs[project_tag] = {}

    project_costs[project_tag][service] = {
        'Current Cost': amount
    }

# Process the previous month data
for item in cost_data_previous:
    project_tag = next((tag.split('$')[1] for tag in item['Keys'] if tag.startswith('Project$')), 'No Project Tag')
    service = next((svc for svc in item['Keys'] if not svc.startswith('Project$')), 'No Service')
    amount = float(item['Metrics']['UnblendedCost']['Amount'])

    if project_tag not in project_costs:
        project_costs[project_tag] = {}

    if service not in project_costs[project_tag]:
        project_costs[project_tag][service] = {}

    project_costs[project_tag][service]['Previous Cost'] = amount

# Output file path
output_file = 'Projectwise_Cost_Report.xlsx'

# First, create the Excel file with data
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    used_sheet_names = set()

    for project, services in project_costs.items():
        sheet_name = get_unique_sheet_name(project, used_sheet_names)
        used_sheet_names.add(sheet_name)

        # Prepare data for the DataFrame
        data = []
        for service, costs in services.items():
            current_cost = costs.get('Current Cost', 0)
            previous_cost = costs.get('Previous Cost', 0)
            difference = current_cost - previous_cost
            data.append({
                'Service': service,
                'Current Cost (USD)': round(current_cost, 2),
                'Previous Cost (USD)': round(previous_cost, 2),
                'Difference (USD)': round(difference, 2)
            })

        df = pd.DataFrame(data)
        df = df.sort_values(by='Current Cost (USD)', ascending=False)

        # Calculate and add totals
        total_current_cost = df['Current Cost (USD)'].sum()
        total_previous_cost = df['Previous Cost (USD)'].sum()
        total_difference = total_current_cost - total_previous_cost

        total_row = pd.DataFrame({
            'Service': ['Total'],
            'Current Cost (USD)': [round(total_current_cost, 2)],
            'Previous Cost (USD)': [round(total_previous_cost, 2)],
            'Difference (USD)': [round(total_difference, 2)]
        })
        df = pd.concat([df, total_row], ignore_index=True)

        # Write to Excel
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)

# Now apply formatting
workbook = load_workbook(output_file)

for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

    # Define colors and borders
    light_yellow = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
    light_cyan = PatternFill(start_color='E0FFFF', end_color='E0FFFF', fill_type='solid')
    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Add title
    title = f"Report Period: {time_period_previous['Start']} to {time_period_previous['End']} & "
    title += f"{time_period_current['Start']} to {time_period_current['End']}"

    sheet.cell(row=1, column=1, value=title)
    sheet.cell(row=1, column=1).fill = light_yellow
    sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=1, column=1).border = thin_border
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # Get the last row of data
    last_row = sheet.max_row

    # Apply formatting to all cells
    for row in sheet.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=4):
        for cell in row:
            if cell.column == 4:  # Difference column
                try:
                    value = float(cell.value or 0)
                    if value > 0:
                        cell.fill = red_fill  # Red for cost exceeded
                    elif value < 0:
                        cell.fill = green_fill  # Green for cost saved
                    else:
                        cell.fill = light_cyan
                except (ValueError, TypeError):
                    cell.fill = light_cyan
            else:
                cell.fill = light_yellow if cell.column == 1 else light_cyan
            cell.border = thin_border

    # Adjust column widths
    for col in range(1, 5):
        sheet.column_dimensions[get_column_letter(col)].width = 30

workbook.save(output_file)
print(f"Consolidated report saved as {output_file}")
