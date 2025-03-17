import boto3
import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta

# Initialize the AWS Cost Explorer client
client = boto3.client('ce', region_name='ca-central-1')

# Function to calculate the date ranges for the two previous months
def get_date_ranges():
    today = datetime.today()
    current_month_start = datetime(today.year, today.month, 1)

    # Calculate previous two months
    first_month_start = current_month_start - relativedelta(months=1)
    first_month_end = current_month_start - timedelta(days=1)
    second_month_start = current_month_start - relativedelta(months=2)
    second_month_end = first_month_start - timedelta(days=1)

    return {
        "first_month": {
            "start_date": second_month_start.strftime('%Y-%m-%d'),
            "end_date": second_month_end.strftime('%Y-%m-%d'),
        },
        "second_month": {
            "start_date": first_month_start.strftime('%Y-%m-%d'),
            "end_date": first_month_end.strftime('%Y-%m-%d'),
        },
    }

# Generate date ranges
date_ranges = get_date_ranges()
first_month = date_ranges["first_month"]
second_month = date_ranges["second_month"]

# Print the calculated time periods for verification
print(f"First Month: {first_month}")
print(f"Second Month: {second_month}")

def fetch_cost_data(time_period):
    # Fetch cost data grouped by the "Project" tag
    response = client.get_cost_and_usage(
        TimePeriod=time_period,
        Granularity='MONTHLY',
        Metrics=['UnblendedCost'],
        GroupBy=[{'Type': 'TAG', 'Key': 'Project'}]
    )

    # Extract and process the cost data
    cost_data = response['ResultsByTime'][0]['Groups']

    # Prepare data for the report
    report_data = []
    for item in cost_data:
        project = next((tag.split('$')[1] for tag in item['Keys'] if tag.startswith('Project$')), 'No Project Tag')
        amount = float(item['Metrics']['UnblendedCost']['Amount'])
        report_data.append({'Project': project, 'Cost (USD)': round(amount, 2)})

    return report_data

# Fetch data for first and second months
first_month_data = fetch_cost_data({
    'Start': first_month['start_date'],
    'End': first_month['end_date']
})
second_month_data = fetch_cost_data({
    'Start': second_month['start_date'],
    'End': second_month['end_date']
})

# Convert data to DataFrame
first_df = pd.DataFrame(first_month_data)
second_df = pd.DataFrame(second_month_data)

# Merge data for both months, keeping projects from the first month
merged_df = pd.merge(first_df, second_df, on='Project', how='left')

# Swap the columns to ensure the second month is first and first month second
merged_df.rename(columns={
    'Cost (USD)_x': f"Cost of {second_month['start_date']} to {second_month['end_date']}",  # Now the first column
    'Cost (USD)_y': f"Cost of {first_month['start_date']} to {first_month['end_date']}"   # Now the second column
}, inplace=True)

# Calculate the difference (Cost saved or exceeded)
merged_df['Difference'] = merged_df[f"Cost of {first_month['start_date']} to {first_month['end_date']}"] - merged_df[f"Cost of {second_month['start_date']} to {second_month['end_date']}"]

# Calculate totals for each column
total_first_month_cost = merged_df[f"Cost of {first_month['start_date']} to {first_month['end_date']}"].sum()
total_second_month_cost = merged_df[f"Cost of {second_month['start_date']} to {second_month['end_date']}"].sum()
total_difference = merged_df['Difference'].sum()

# Create a new row for totals
totals_row = pd.DataFrame({
    'Project': ['Total'],
    f"Cost of {first_month['start_date']} to {first_month['end_date']}": [total_first_month_cost],
    f"Cost of {second_month['start_date']} to {second_month['end_date']}": [total_second_month_cost],
    'Difference': [total_difference]
})

# Append totals row to the DataFrame
merged_df = pd.concat([merged_df, totals_row], ignore_index=True)

# Save data to Excel with merged costs and the Difference column
output_file = 'AWS_Project_Cost_Report_with_Difference_and_Totals.xlsx'

# Create an Excel writer
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    # Write the merged data to the Excel sheet
    merged_df.to_excel(writer, index=False, sheet_name='Report', startrow=2, startcol=0)

# Add date range title and formatting
wb = load_workbook(output_file)
sheet = wb['Report']

# Define colors and borders
first_month_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')  # Light Yellow
second_month_fill = PatternFill(start_color='E0FFFF', end_color='E0FFFF', fill_type='solid')  # Light Cyan
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Insert date range title
sheet.cell(row=1, column=1).value = f"Report Period: {first_month['start_date']} to {first_month['end_date']} & {second_month['start_date']} to {second_month['end_date']}"
sheet.cell(row=1, column=1).fill = first_month_fill
sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
sheet.cell(row=1, column=1).border = thin_border
sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

# Apply color and border formatting for the table
# Ensure that the range for iter_rows is correctly set
last_row = 2 + len(merged_df)  # Update the max_row calculation

for row in sheet.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=4):
    for cell in row:
        # Check if the value in the "Difference" column is numeric before comparison
        if cell.column == 4:  # Difference column
            try:
                difference = float(cell.value)  # Convert to float
            except (ValueError, TypeError):
                difference = 0  # Default to 0 if the value is invalid or empty

            # Color the "Difference" column based on the value
            if difference > 0:
                cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')  # Red for cost exceeded
            elif difference < 0:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green for cost saved
            else:
                cell.fill = second_month_fill  # Light Cyan for no change
        else:
            cell.fill = first_month_fill if cell.column == 1 else second_month_fill  # Color projects and costs
        cell.border = thin_border

# Adjust column widths for better readability
for col in range(1, 5):  # Adjust columns A to D (Including Difference)
    column_letter = get_column_letter(col)
    sheet.column_dimensions[column_letter].width = 30

# Save the updated workbook
wb.save(output_file)

# Print confirmation
print(f'Report saved as {output_file}')
